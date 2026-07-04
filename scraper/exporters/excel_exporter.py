"""Business-ready Excel workbook export.

Every run produces ``exports/moneycontrol_it_YYYYMMDD_HHMMSS.xlsx`` with one
sheet per data family. Formatting applied to every sheet:

* bold white header row on a dark-blue fill,
* frozen header row and auto-filter,
* column widths sized to content (capped, long text wrapped),
* Indian-style thousands separators for numeric columns,
* ISO date formatting for timestamp columns.
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from scraper.models.company import RunStats

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
CELL_ALIGN = Alignment(vertical="top", wrap_text=False)
WRAP_ALIGN = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))

NUMBER_FORMAT = "#,##0.00"
INT_FORMAT = "#,##0"

# Columns considered "long text": get wrapping + a wide fixed width.
_WRAP_WIDTH = 60
_MAX_WIDTH = 45


class ExcelExporter:
    """Writes the multi-sheet workbook for one run."""

    def __init__(self, export_dir: Path) -> None:
        self._export_dir = export_dir

    # -- generic sheet writer ------------------------------------------------------

    @staticmethod
    def _write_sheet(
        ws: Worksheet,
        headers: list[str],
        rows: list[list[object]],
        wrap_columns: set[str] | None = None,
        number_columns: set[str] | None = None,
    ) -> None:
        wrap_columns = wrap_columns or set()
        number_columns = number_columns or set()

        ws.append(headers)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGN

        for row in rows:
            ws.append(row)

        widths = [len(h) for h in headers]
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                header = headers[col_idx - 1]
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = THIN_BORDER
                if header in wrap_columns:
                    cell.alignment = WRAP_ALIGN
                else:
                    cell.alignment = CELL_ALIGN
                if isinstance(value, float) and header in number_columns:
                    cell.number_format = (
                        INT_FORMAT if value == int(value) else NUMBER_FORMAT
                    )
                if value is not None and header not in wrap_columns:
                    widths[col_idx - 1] = max(
                        widths[col_idx - 1], min(len(str(value)), _MAX_WIDTH)
                    )

        for col_idx, header in enumerate(headers, start=1):
            letter = get_column_letter(col_idx)
            if header in wrap_columns:
                ws.column_dimensions[letter].width = _WRAP_WIDTH
            else:
                ws.column_dimensions[letter].width = widths[col_idx - 1] + 3

        ws.freeze_panes = "A2"
        if rows:
            ws.auto_filter.ref = (
                f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"
            )

    # -- workbook -----------------------------------------------------------------

    def export(
        self,
        companies: list[dict],
        index_membership: dict[str, list[str]],
        quarterly: list[dict],
        annual: list[dict],
        shareholding: list[dict],
        stats: RunStats,
    ) -> Path:
        """Write the workbook and return its path."""
        self._export_dir.mkdir(parents=True, exist_ok=True)
        path = self._export_dir / (
            f"moneycontrol_it_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        )
        wb = Workbook()

        self._sheet_company_list(wb.active, companies, index_membership)
        self._sheet_company_profile(wb.create_sheet("Company Profile"), companies)
        self._sheet_financial_data(wb.create_sheet("Financial Data"), companies, annual)
        self._sheet_periods(wb.create_sheet("Quarterly Results"), quarterly)
        self._sheet_shareholding(wb.create_sheet("Shareholding"), shareholding)
        self._sheet_facet(
            wb.create_sheet("Products & Services"), companies,
            "products_services", "Products & Services",
        )
        self._sheet_facet(
            wb.create_sheet("Clients"), companies, "major_clients", "Major Clients"
        )
        self._sheet_facet(
            wb.create_sheet("Order Book"), companies, "order_book", "Order Book"
        )
        self._sheet_errors(wb.create_sheet("Errors & Skipped"), stats)
        self._sheet_summary(wb.create_sheet("Execution Summary"), stats)

        wb.save(path)
        logger.info("Excel workbook written: %s", path)
        return path

    # -- individual sheets -----------------------------------------------------------

    def _sheet_company_list(
        self, ws: Worksheet, companies: list[dict],
        index_membership: dict[str, list[str]],
    ) -> None:
        ws.title = "Company List"
        headers = ["Company Name", "NSE Symbol", "BSE Symbol", "ISIN", "Sector",
                   "Industry", "Market Cap (Rs Cr)", "NSE Indices",
                   "Moneycontrol URL"]
        rows = [
            [c.get("company_name"), c.get("nse_symbol"), c.get("bse_symbol"),
             c.get("isin"), c.get("sector"), c.get("industry"),
             c.get("market_cap_cr"),
             ", ".join(index_membership.get(c.get("nse_symbol", ""), [])),
             c.get("mc_url")]
            for c in companies
        ]
        self._write_sheet(ws, headers, rows,
                          number_columns={"Market Cap (Rs Cr)"})

    def _sheet_company_profile(self, ws: Worksheet, companies: list[dict]) -> None:
        headers = ["Company Name", "NSE Symbol", "Sector", "Industry",
                   "Business Summary", "Website", "Email", "Phone", "Address",
                   "Registrar", "Management", "Static Updated (UTC)",
                   "Dynamic Updated (UTC)"]
        rows = [
            [c.get("company_name"), c.get("nse_symbol"), c.get("sector"),
             c.get("industry"), c.get("business_summary"), c.get("website"),
             c.get("email"), c.get("phone"), c.get("address"),
             c.get("registrar"), c.get("management"),
             c.get("static_updated_at"), c.get("dynamic_updated_at")]
            for c in companies
        ]
        self._write_sheet(
            ws, headers, rows,
            wrap_columns={"Business Summary", "Management", "Address"},
        )

    def _sheet_financial_data(
        self, ws: Worksheet, companies: list[dict], annual: list[dict]
    ) -> None:
        """Market snapshot + latest annual revenue/profit per company."""
        latest_annual: dict[str, dict] = {}
        for row in annual:  # rows arrive newest-first per company; keep first
            latest_annual.setdefault(row["nse_symbol"], row)
        headers = ["Company Name", "NSE Symbol", "Market Cap (Rs Cr)", "Price",
                   "P/E", "P/B", "Industry P/E", "Dividend Yield %",
                   "Book Value", "EPS (TTM)", "Face Value", "52W High",
                   "52W Low", "Latest FY", "FY Revenue (Rs Cr)",
                   "FY Net Profit (Rs Cr)"]
        number_cols = set(headers[2:]) - {"Latest FY"}
        rows = []
        for c in companies:
            fy = latest_annual.get(c.get("nse_symbol", ""), {})
            rows.append(
                [c.get("company_name"), c.get("nse_symbol"),
                 c.get("market_cap_cr"), c.get("price"), c.get("pe"),
                 c.get("pb"), c.get("industry_pe"), c.get("dividend_yield"),
                 c.get("book_value"), c.get("eps_ttm"), c.get("face_value"),
                 c.get("week52_high"), c.get("week52_low"),
                 fy.get("period_label"), fy.get("revenue"),
                 fy.get("net_profit")]
            )
        self._write_sheet(ws, headers, rows, number_columns=number_cols)

    def _sheet_periods(self, ws: Worksheet, periods: list[dict]) -> None:
        headers = ["Company Name", "NSE Symbol", "Period", "Revenue (Rs Cr)",
                   "Other Income (Rs Cr)", "Total Income (Rs Cr)",
                   "Expenditure (Rs Cr)", "Interest (Rs Cr)", "Tax (Rs Cr)",
                   "Net Profit (Rs Cr)", "Basic EPS"]
        rows = [
            [p.get("company_name"), p.get("nse_symbol"), p.get("period_label"),
             p.get("revenue"), p.get("other_income"), p.get("total_income"),
             p.get("expenditure"), p.get("interest"), p.get("tax"),
             p.get("net_profit"), p.get("basic_eps")]
            for p in periods
        ]
        self._write_sheet(ws, headers, rows, number_columns=set(headers[3:]))

    def _sheet_shareholding(self, ws: Worksheet, shareholding: list[dict]) -> None:
        headers = ["Company Name", "NSE Symbol", "Category", "Holding %",
                   "Updated (UTC)"]
        rows = [
            [s.get("company_name"), s.get("nse_symbol"), s.get("category"),
             s.get("percent"), s.get("updated_at")]
            for s in shareholding
        ]
        self._write_sheet(ws, headers, rows, number_columns={"Holding %"})

    def _sheet_facet(
        self, ws: Worksheet, companies: list[dict], key: str, label: str
    ) -> None:
        headers = ["Company Name", "NSE Symbol", label]
        rows = [
            [c.get("company_name"), c.get("nse_symbol"), c.get(key)]
            for c in companies
            if c.get(key)
        ]
        if not rows:
            rows = [[f"No {label.lower()} information found in company profiles",
                     None, None]]
        self._write_sheet(ws, headers, rows, wrap_columns={label})

    def _sheet_errors(self, ws: Worksheet, stats: RunStats) -> None:
        headers = ["NSE Symbol", "Company Name", "Stage", "Error",
                   "Occurred At"]
        rows = [
            [e.nse_symbol, e.company_name, e.stage, e.message,
             e.occurred_at.strftime("%Y-%m-%d %H:%M:%S")]
            for e in stats.errors
        ]
        missing_rows = [
            [symbol, None, "missing-fields", ", ".join(fields), None]
            for symbol, fields in sorted(stats.missing_fields.items())
        ]
        self._write_sheet(ws, headers, rows + missing_rows,
                          wrap_columns={"Error"})

    def _sheet_summary(self, ws: Worksheet, stats: RunStats) -> None:
        headers = ["Metric", "Value"]
        finished = stats.finished_at or datetime.now()
        rows: list[list[object]] = [
            ["Run started", stats.started_at.strftime("%Y-%m-%d %H:%M:%S")],
            ["Run finished", finished.strftime("%Y-%m-%d %H:%M:%S")],
            ["Duration (seconds)", round(stats.duration_seconds, 1)],
            ["Indices fetched", stats.indices_fetched],
            ["Constituents seen (all sectors)", stats.constituents_seen],
            ["IT companies identified", stats.it_companies],
            ["Companies scraped successfully", stats.companies_scraped],
            ["Static profiles refreshed", stats.static_refreshed],
            ["Static cache hits (skipped)", stats.static_cache_hits],
            ["Companies failed", stats.companies_failed],
            ["HTTP requests made", stats.requests_made],
            ["Errors recorded", len(stats.errors)],
            ["Companies with missing fields", len(stats.missing_fields)],
        ]
        self._write_sheet(ws, headers, rows)
